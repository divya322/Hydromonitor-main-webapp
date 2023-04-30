# import pandas lib as pd
import datetime
import pytz
import pandas as pd
from os.path import exists
import openpyxl

def get_azure_data(data):
    water,light,soil,absoil,temp,humidity,npk,flap=data.split(" ")
    ct = str(datetime.datetime.now(pytz.timezone('Asia/Kolkata')))
    ct = ct[11:16]
    h=int(ct[0:2])
    if (h>12):
        ct = ct[3:]
        h-=12
        t="PM"
    else:
        ct = ct[3:]
        t="AM"
    time=str(h)+":"+ct+" "+t
    print(time)
    row = [temp,humidity,light,water,soil,absoil,npk,flap,time]
    return row

def len_of_excel():     
        
    if (exists('data.xlsx')==False):
        import openpyxl
        wb = openpyxl.Workbook() 
        sheet = wb['Sheet']
        sheet.title = 'Sheet1'
        row=1
        new_row = ['temp','humidity','light','soil','absoil','water','npk','flap','time']
        for col, entry in enumerate(new_row, start=1):
            sheet.cell(row=row, column=col, value=entry)
        wb.save(filename='data.xlsx')
        
    dataframe1 = pd.read_excel('data.xlsx')
    len_excel=len(dataframe1['time'])
    return len_excel

def write_to_excel(len_excel,new_row):

    workbook = openpyxl.load_workbook(filename='data.xlsx')
    worksheet = workbook['Sheet1'] 
    
    row=len_excel+2
    for col, entry in enumerate(new_row, start=1):
        worksheet.cell(row=row, column=col, value=entry)

    workbook.save('data.xlsx')

def getdata(cl_name):
    if (exists('data.xlsx')==False):
        len_of_excel()
    # read by default 1st sheet of an excel file
    dataframe1 = pd.read_excel('data.xlsx')
    cl_data = list(dataframe1[cl_name])
    if (cl_data==[] and cl_name=="time"):
        data="xx:xx xx"
    if (cl_data==[] and cl_name!="time"):
        data="xx"
    if(cl_data!=[]):
        data=cl_data[-1]
    return data


